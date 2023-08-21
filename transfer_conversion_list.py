import datetime
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import xlrd
import xlwings as xw


# -*- coding = utf-8 -*-
# @Time : 2023/4/25 下午 9:12
# @Author : Randy
# @File : transfer_conversion_list.py
# @Software : PyCharm

def transfer_conversion(input_file: str, input_sheet: str, conversion_list_name: str,
                        anno_sheet: str, anno_file: str, temp_file: str,
                        temp_sheet_column: str, temp_quad: str):
    '''
    :param input_file: copy the second row from the annotation file and then
    transpose the row to a column and paste to the file
    :param input_sheet: the new file's sheet name
    :param conversion_list_name: the output conversion list file's name
    :param anno_sheet: the sheet name we use in the annotation file
    :param anno_file: the annotation file's name
    :param temp_file:  template file's name
    :param temp_sheet_column: template file's sheet and column start column form should be 'Sheet7!M'
    :param temp_quad: the position of columns added to the file, format should be 'Sheet7!K2'
    :return:
    '''
    wb = load_workbook(input_file)
    ws = wb.get_sheet_by_name(input_sheet)
    row_num = 1
    last_photo_name = ''
    # set the useful value start from B column
    xlsx_file = Workbook()
    xlsx_ws = xlsx_file.active
    start_column = 0
    for row in ws.rows:
        start_column += 1
        photo_name = row[0].value[0: 12]
        # Start with HRI, HRO, SHI, SHO
        if photo_name[0:2] == 'HR' or photo_name[0:2] == 'SH':
            if photo_name[-1].isdigit():
                photo_name = photo_name[:11]
        # start with DB
        elif photo_name[0:2] == 'DB':
            photo_name = photo_name[0:11]
            if photo_name[-1].isdigit():
                photo_name = photo_name[:10]
        # not a photo name
        else:
            continue
        if photo_name != last_photo_name:
            xlsx_ws['A' + str(row_num)] = anno_file
            xlsx_ws['B' + str(row_num)] = anno_sheet + '!' + get_column_letter(start_column)
            # I will count the photo's columns' number later (column C)
            xlsx_ws['D' + str(row_num)] = temp_file
            xlsx_ws['E' + str(row_num)] = temp_sheet_column
            xlsx_ws['F' + str(row_num)] = photo_name + '.xlsx'
            # offset
            xlsx_ws['G' + str(row_num)] = 0
            xlsx_ws['H' + str(row_num)] = temp_quad
            row_num += 1
            last_photo_name = photo_name
    xlsx_file.save(conversion_list_name)
    wb.close()


def transfer_old_conversion(input_file: str, input_sheet: str,
                            anno_sheet: str, anno_file: str, temp_file: str,
                            temp_sheet_column: str, temp_quad: str, prefix: str, suffix: str):
    '''
    :param input_file: copy the second row from the annotation file and then
    transpose the row to a column and paste to the file
    :param input_sheet: the new file's sheet name
    :param anno_sheet: the sheet name we use in the annotation file
    :param anno_file: the annotation file's name
    :param temp_file:  template file's name
    :param temp_sheet_column: template file's sheet and column start column form should be 'Sheet7!M'
    :param temp_quad: the position of columns added to the file, format should be 'Sheet7!K2'
    :return:
    '''
    wb = load_workbook(input_file)
    ws = wb.get_sheet_by_name(input_sheet)
    row_num = 1
    # set the useful value start from B column
    old_date = ''
    xlsx_file = Workbook()
    xlsx_ws = xlsx_file.active
    start_column = 0
    for row in ws.rows:
        start_column += 1
        date = row[0].value
        if not isinstance(date, datetime.datetime):
            continue
        if date != old_date:
            transferred_date = datetime.datetime.strftime(date, "%m%d%Y")
            xlsx_ws['A' + str(row_num)] = anno_file[0: anno_file.find('_')] + '.xls'
            xlsx_ws['B' + str(row_num)] = anno_sheet + '!' + get_column_letter(start_column)
            # I will count the photo's columns' number later (column C)
            xlsx_ws['D' + str(row_num)] = temp_file
            xlsx_ws['E' + str(row_num)] = temp_sheet_column
            xlsx_ws['F' + str(row_num)] = prefix + transferred_date[0:4] + \
                                          transferred_date[6:] + suffix + '.xlsx'
            # offset
            xlsx_ws['G' + str(row_num)] = 0
            xlsx_ws['H' + str(row_num)] = temp_quad
            row_num += 1
            old_date = date
    xlsx_file.save(anno_file[:anno_file.find('.')] + '_ConversionList.xlsx')
    wb.close()
